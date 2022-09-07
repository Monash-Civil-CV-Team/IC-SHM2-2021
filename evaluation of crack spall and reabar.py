from model_p2_234 import *
# from segnet import *
# from model_self_8115_p5 import *
import os.path
from cal_resize_iou import cal_resize_iou
import cv2 as cv
import imageio
import numpy as np
import shutil
from skimage import transform as trans
#from model_self_85 import *
from PIL import Image
import openpyxl as xl
#from model_att_unet import *
from model_att_unet import *
from model_att_unet_2 import *

os.environ["CUDA_VISIBLE_DEVICES"] = "0"

# os.environ['TF_FORCE_GPU_ALLOW_GROWTH'] = 'true'

fullscale_img_dir = 'D://IPC_SHM2//image_test//'
#fullscale_img_dir = 'D://IPC_SHM2//project2//fullsize//crack_image//'
# fullscale_label_dir = 'project2//fullsize//crack//'

# --------------------------------------------------Test----------------------------------------------------


def crop(source_dir, crop_img_dir):
    for file in os.listdir(source_dir):
        img = imageio.imread(source_dir + file)
        img = cv.resize(img, (4 * 512, 2 * 512), cv.INTER_NEAREST)
        img = np.array(img)
        for row in range(2):
            for col in range(4):
                cropped = img[512 * row:512 * (row + 1), 512 * col:512 * (col + 1), :]
                imageio.imwrite(crop_img_dir + file.split('.')[0] + '%d%d.png' % (row, col), cropped)


def batch_test(model_name, model, test_input_dir,test_save_dir, colored_test_save_dir):
#def batch_test(model_name, model, test_input_dir,test_save_dir):
    os.makedirs(colored_test_save_dir, exist_ok=True)
    os.makedirs(test_save_dir, exist_ok=True)
    model.load_weights(model_name + '.hdf5')
    data_name = os.listdir(test_input_dir)
    threshold = 0.5
    for i in data_name:
        img = imageio.imread(os.path.join(test_input_dir, i), as_gray=False)
        img = img/255
        # img = trans.resize(img, (512, 512), mode='edge')
        img = np.reshape(img, (1,) + img.shape)
        results = model.predict(img, verbose=1)
        results = np.squeeze(results)
        results[results <= threshold] = 0
        results[results > threshold] = 1

        filename = i.split('.')[0]
        # imageio.imsave(os.path.join(colored_test_save_dir, filename + ".png"), results)
        results = results.astype(np.uint8)
        imageio.imsave(os.path.join(test_save_dir, filename + ".png"), results)
        imageio.imsave(os.path.join(colored_test_save_dir, filename + ".png"), results*255)


def batch_test_new(model_name, model, test_input_dir,test_save_dir):
    """This is to save the predictions in numpy array format without any thresholding.
    The savings will be merged later on where the thresholding will be applied."""
    os.makedirs(test_save_dir, exist_ok=True)
    model.load_weights(model_name + '.hdf5')
    data_name = os.listdir(test_input_dir)
    threshold = 0.5
    for i in data_name:
        img = imageio.imread(os.path.join(test_input_dir, i), as_gray=False)
        img = img/255
        # img = trans.resize(img, (512, 512), mode='edge')
        img = np.reshape(img, (1,) + img.shape)
        results = model.predict(img, verbose=1)
        results = np.squeeze(results)
        # results[results <= threshold] = 0
        # results[results > threshold] = 1

        filename = i.split('.')[0]

        # results = results.astype(np.uint8)
        np.save(os.path.join(test_save_dir, filename + ".npy"), results)


def joint_resize(crop_input,save_dir, if_paint=False):
    # os.makedirs(colored_save_dir, exist_ok=True)
    os.makedirs(save_dir, exist_ok=True)
    for i in range(0, len(os.listdir(crop_input)), 8):
        merge = np.zeros((1024, 2048))
        for j in range(i, i + 8):
            file_name = os.listdir(crop_input)[j]
            crop = imageio.imread(crop_input+file_name, as_gray=True)
            row = file_name.split('.')[0][-2]
            col = file_name.split('.')[0][-1]
            merge[int(row) * 512:(int(row) + 1) * 512, int(col) * 512:(int(col) + 1) * 512] = crop
        merge = cv.resize(merge, (1920, 1080), cv.INTER_NEAREST)
        merge = np.round(merge)
        # imageio.imwrite(colored_save_dir + os.listdir(crop_input)[i].split('_')[0][0: 6] + '.png', merge)
        if not if_paint:
            merge = merge.astype(np.uint8)
        imageio.imwrite(save_dir + os.listdir(crop_input)[i][0: 6] + '.png', merge)


def joint_resize_new(crop_input, save_dir, if_paint=False):
    # os.makedirs(colored_save_dir, exist_ok=True)
    os.makedirs(save_dir, exist_ok=True)
    for i in range(0, len(os.listdir(crop_input)), 8):
        merge = np.zeros((1024, 2048))
        for j in range(i, i + 8):
            file_name = os.listdir(crop_input)[j]
            crop = np.load(crop_input+file_name)
            row = file_name.split('.')[0][-2]
            col = file_name.split('.')[0][-1]
            merge[int(row) * 512:(int(row) + 1) * 512, int(col) * 512:(int(col) + 1) * 512] = crop
        merge = cv.resize(merge, (1920, 1080), cv.INTER_NEAREST)
        merge = np.round(merge)
        # imageio.imwrite(colored_save_dir + os.listdir(crop_input)[i].split('_')[0][0: 6] + '.png', merge)
        if not if_paint:
            merge = merge.astype(np.uint8)
        imageio.imwrite(save_dir + os.listdir(crop_input)[i][0: 6] + '.png', merge)


def confusion_matrix(label, pred, n_class=2): #用np.bincount（）计算混淆矩阵
    mask = (label >= 0) & (label < n_class)
    hist = np.bincount(
        n_class*label[mask].astype(int) + pred[mask],
        minlength=n_class ** 2).reshape(n_class, n_class)
    return hist


def indicators(pred_dir, gt_dir):
    """Calculate the confusion matrix and the accuracy criteria between the prediction and ground truth in batch."""

    cm = {'file': [], 'TP': [], 'FP': [], 'TN': [], 'FN': []}
    indics = {'file': [], 'iou': [], 'accuracy': [], 'precision': [], 'recall': [], 'f1': []}
    for file in os.listdir(pred_dir):
        pred = Image.open(pred_dir + file)
        pred = np.array(pred)
        gt = Image.open(gt_dir + file)
        gt = np.array(gt)
        hist = confusion_matrix(gt, pred)
        TP, FP, TN, FN = hist[1, 1], hist[0, 1], hist[0, 0], hist[1, 0]
        cm['file'].append(file)
        cm['TP'].append(TP)
        cm['FP'].append(FP)
        cm['TN'].append(TN)
        cm['FN'].append(FN)
        iou_foreground = TP / (TP+FP+FN)
        acc = (TP + TN) / (TP + FP + TN + FN)
        pre = TP / (TP + FP)
        recall = TP / (TP + FN)
        F1 = 2 * (acc * recall) / (pre + recall)
        indics['file'].append(file)
        indics['iou'].append(iou_foreground)
        indics['accuracy'].append(acc)
        indics['precision'].append(pre)
        indics['recall'].append(recall)
        indics['f1'].append(F1)
    cm['file'].append('mean')
    cm['TP'].append(np.mean(cm['TP']))
    cm['FP'].append(np.mean(cm['FP']))
    cm['TN'].append(np.mean(cm['TN']))
    cm['FN'].append(np.mean(cm['FN']))
    indics['file'].append('mean')
    indics['iou'].append(np.mean(indics['iou']))
    indics['accuracy'].append(np.mean(indics['accuracy']))
    indics['precision'].append(np.mean(indics['precision']))
    indics['recall'].append(np.mean(indics['recall']))
    indics['f1'].append(np.mean(indics['f1']))
    return cm, indics


def eval(source_dir, save_dir, model_name, model, label_dir =None):
    """
    This is to enable end-to-end evaluation of full-scale images with size of 1960*1080, save predicted masks and
    output indicators including IoU, accuracy, precision, recall, and F1 score.
    :param
    source_dir: the directory of the test fullscale images
    save_dir: the root save path
    model_name: weight to be loaded
    model: structure of the model to be tested
    :return
    The indicator array of [IoU, accuracy, precision, recall, F1 score]
    Predicted masks, which are saved within the root save path.
    """
    save_dir = save_dir + model_name + '/'
    os.makedirs(save_dir, exist_ok=True)
    # crop_img_dir = save_dir+'/crop_img/'
    # os.makedirs(crop_img_dir, exist_ok=True)

    #crop(source_dir, crop_img_dir)
    crop_img_dir = 'project2/test results/crop_img_8/'
    #crop_img_dir = 'project2/test results/crop_img/'
    #crop_img_dir = 'project2/3724crack/image1/'
    batch_test(model_name, model, crop_img_dir, save_dir+'crop_mask/',save_dir+'colored_crop_mask/')

    joint_resize(save_dir + 'crop_mask/', save_dir + 'full_mask/', if_paint=False)
    joint_resize(save_dir+'crop_mask/', save_dir+'colored_full_mask/', if_paint=True)


def eval_new(source_dir, save_dir, model_name, model, label_dir =None):
    save_dir = save_dir + model_name + '/'
    os.makedirs(save_dir, exist_ok=True)

    crop_img_dir = 'project2/test results/crop_img_8/'

    batch_test_new(model_name, model, crop_img_dir, save_dir+'crop_mask/')

    joint_resize_new(save_dir + 'crop_mask/', save_dir + 'full_mask/', if_paint=False)
    joint_resize_new(save_dir+'crop_mask/', save_dir+'colored_full_mask/', if_paint=True)
    shutil.rmtree(save_dir+'crop_mask/')


def write_iou_xlsx(path, sheet_name, value): #iou结果写入excel文件
    index = len(value)
    workbook = xl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
    workbook.save(path)
    print('xlsx表格写入数据成功')


def eval_to_cm(img_in, save_dir, model, weight, gt_crop, gt_full):

    """This is to conduct inference on the labelled test set and save confusion matrices and
    indicators including IoU, accuracy, precision, recall, and F1 score in xlsx files."""

    #make predictions
    save_dir = save_dir + weight + '/val/'
    colored_mask = save_dir+'colored_crop_mask/'
    crop_mask = save_dir+'crop_mask/'
    batch_test(weight, model, img_in, crop_mask, colored_mask)

    full_mask = save_dir+'full_mask/'
    colored_full_mask = save_dir+'colored_full_mask/'
    os.makedirs(full_mask, exist_ok=True)
    os.makedirs(colored_full_mask, exist_ok=True)
    joint_resize(crop_mask, full_mask, if_paint=False)
    joint_resize(colored_mask, colored_full_mask, if_paint=True)

    #calculate indicators on crop and full images
    cm_crop, indics_crop = indicators(crop_mask, gt_crop)
    cm_full, indics_full = indicators(full_mask, gt_full)
    workbook = xl.Workbook()
    sheet = workbook.active
    sheet.title = 'try'
    titles = ['file', 'TP', 'FP', 'TN', 'FN', 'iou', 'accuracy', 'precision', 'recall', 'f1']
    for i in range(len(titles)):
        sheet.cell(row=1, column=i + 1, value=titles[i])
    for i in range(len(cm_full['file'])):
        for j in range(5):
            sheet.cell(row=i + 2, column=j + 1, value=cm_full[titles[j]][i])
        for j in range(5, 10):
            sheet.cell(row=i + 2, column=j + 1, value=indics_full[titles[j]][i])
    workbook.save(save_dir +weight + '.xlsx')
    print('xlsx表格写入数据成功')


if __name__ == '__main__':
       eval(fullscale_img_dir, save_dir='project2/test results/',
                 model_name='newdata_crack_lowthre_unet_bs4', model=unet())
      # eval_to_cm('D:/IPC_SHM2/project2/new_data/crack/val/img/', 'project2/test results/', unet(),
       #            'newdata_crack_lowthre_unet_bs4', 'D:/IPC_SHM2/project2/new_data/crack/val/label/',
       #            'D:/IPC_SHM2/project2/new_data/crack/val/full_label/')
      # model_name = 'newdata_crack_attunet_v9_newres', model = Attention_ResUNet())
      #        #model_name='linknet_t2_crackv3', model=LinkNet())
      #       # model_name = 'unet_t4_rebarv3', model = unet())
      #      #model_name='newdata_crack_crackdet_v2_lr5e4', model=Crackdet())
      #  # model_name = 'new_data_crack_unet_v1', model = unet())
      #       #   model_name = 'unet_t2_crackv3', model = unet())
      # #   model_name = 'newdata_crack_unet_v2_lr1e5', model = unet())
      #       # model_name = 'Crackdet_t2_crackv3', model = Crackdet())
      #        model_name = 'selfnet_p2_85V3', model = newmodel_noname())
          #   model_name = 'Crackdet_t2_crackv3', model = Crackdet())
    # joint_resize('project2//test results//unet_t3_spallv3//crop_mask//', 'project2//test results//unet_t3_spallv3//full_mask//')
    # cal_resize_iou('project2//test results//Crackdet_t2_crackv1//full_mask//', fullscale_label_dir, 'project2/test results/', model_name='Crackdet_t2_crackv1')
    # batch_test('Crackdet_t2_crackv3', Crackdet(), 'project2//test results//crop_img//', 'project2//test results//Crackdet_t2_crackv3//crop_mask//')
  #         #
  # eval_to_cm('D:/IPC_SHM2/project2/new_data/crack/val/img/', 'project2/val_results/', Attention_ResUNet(),
  #            'newdata_crack_attunet_v9_newres', 'D:/IPC_SHM2/project2/new_data/crack/val/label/',
  #            'D:/IPC_SHM2/project2/new_data/crack/val/full_label/')
